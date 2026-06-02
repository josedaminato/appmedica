from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import bad_request

MAX_ROWS = 2000
MAX_FILE_BYTES = 5 * 1024 * 1024


def parse_spreadsheet(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    if len(file_bytes) > MAX_FILE_BYTES:
        raise bad_request("El archivo supera el límite de 5 MB")

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise bad_request("No se pudo leer el archivo Excel (.xlsx)") from exc

    sheet = workbook.active
    if sheet is None:
        raise bad_request("El archivo no tiene hojas")

    rows_iter = sheet.iter_rows(values_only=True)
    header_row: list[str] | None = None
    data_rows: list[dict[str, Any]] = []

    for excel_row in rows_iter:
        cells = [_cell_to_str(c) for c in excel_row]
        if not any(cells):
            continue

        if header_row is None:
            header_row = _normalize_headers(cells)
            continue

        if len(data_rows) >= MAX_ROWS:
            break

        row_dict: dict[str, Any] = {}
        for idx, header in enumerate(header_row):
            if not header:
                continue
            value = cells[idx] if idx < len(cells) else ""
            row_dict[header] = value.strip() if isinstance(value, str) else value
        if any(str(v).strip() for v in row_dict.values() if v is not None):
            data_rows.append(row_dict)

    workbook.close()

    if not header_row:
        raise bad_request("No se encontraron encabezados en la primera fila con datos")

    return header_row, data_rows


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    return str(value).strip()


def _normalize_headers(cells: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for cell in cells:
        base = cell.strip() or "columna"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers
